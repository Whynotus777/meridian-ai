"""Excel workbook export for CIM analysis results.

Produces a professional multi-tab Excel workbook from an AnalysisResult.
Tabs: Financial Summary | Revenue Breakdown | Comp Set | Risk Register | Deal Score

Requires: openpyxl
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING, Optional
import json
import re

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from core.pipeline import AnalysisResult


# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
_NAVY      = "1F4E79"
_DARK_BLUE = "2E75B6"
_LIGHT_BLUE = "BDD7EE"
_WHITE     = "FFFFFF"
_LIGHT_GREY = "F2F2F2"
_MED_GREY  = "D9D9D9"
_GREEN     = "70AD47"
_AMBER     = "FFC000"
_RED       = "FF0000"
_DARK_RED  = "C00000"
_ORANGE    = "ED7D31"


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_cell(ws, row: int, col: int, value, width_hint: Optional[int] = None):
    """Write a header cell (dark blue background, white bold text)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()
    if width_hint:
        ws.column_dimensions[get_column_letter(col)].width = width_hint
    return cell


def _subheader_cell(ws, row: int, col: int, value):
    """Light-blue sub-header."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_NAVY, size=10)
    cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _thin_border()
    return cell


def _data_cell(ws, row: int, col: int, value, alt_row: bool = False,
               number_format: str = "General", align: str = "left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=_LIGHT_GREY if alt_row else _WHITE)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin_border()
    cell.number_format = number_format
    return cell


def _title_row(ws, title: str, col_span: int):
    """Merge cells across first row for a sheet title."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, color=_WHITE, size=14)
    cell.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _severity_fill(severity: str) -> PatternFill:
    colours = {
        "Critical": _DARK_RED,
        "High":     _RED,
        "Medium":   _AMBER,
        "Low":      _GREEN,
    }
    return PatternFill("solid", fgColor=colours.get(severity, _MED_GREY))


def _score_fill(score: float) -> PatternFill:
    if score >= 0.75:
        return PatternFill("solid", fgColor=_GREEN)
    if score >= 0.50:
        return PatternFill("solid", fgColor=_AMBER)
    return PatternFill("solid", fgColor=_ORANGE)


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def _build_financial_summary(ws, extracted_data: dict, company_name: str):
    fin = extracted_data.get("financials", {})
    rev = fin.get("revenue", {})
    ebitda = fin.get("ebitda", {})
    customers = extracted_data.get("customers", {})

    _title_row(ws, f"Financial Summary — {company_name}", 4)
    ws.row_dimensions[1].height = 30

    # Sub-header row
    for col, header in enumerate(
        ["Category", "Metric", "Value", "Notes"], start=1
    ):
        _header_cell(ws, 2, col, header,
                     width_hint=[22, 32, 20, 40][col - 1])

    rows = [
        ("Revenue",     "LTM Revenue",              _fmt_num(rev.get("ltm")),            "Latest twelve months"),
        ("Revenue",     "Prior Year Revenue",        _fmt_num(rev.get("prior_year")),     ""),
        ("Revenue",     "Two Years Ago Revenue",     _fmt_num(rev.get("two_years_ago")),  ""),
        ("Revenue",     "3-Year CAGR",               _fmt_pct(rev.get("cagr_3yr")),       ""),
        ("Profitability","LTM EBITDA",               _fmt_num(ebitda.get("ltm")),         "Reported"),
        ("Profitability","Adjusted EBITDA (LTM)",    _fmt_num(ebitda.get("adjusted_ebitda_ltm")), "Incl. add-backs"),
        ("Profitability","EBITDA Margin (LTM)",      _fmt_pct(ebitda.get("margin_ltm")),  ""),
        ("Profitability","Gross Margin",             _fmt_pct(fin.get("gross_margin")),   ""),
        ("Profitability","Net Income (LTM)",         _fmt_num(fin.get("net_income")),     ""),
        ("Balance Sheet","Total Debt",               _fmt_num(fin.get("debt")),           ""),
        ("Balance Sheet","Cash & Equivalents",       _fmt_num(fin.get("cash")),           ""),
        ("Balance Sheet","CapEx (Annual)",           _fmt_num(fin.get("capex")),          ""),
        ("Revenue Mix",  "Recurring Revenue %",      _fmt_pct(fin.get("recurring_revenue_pct")), ""),
        ("Customers",   "Total Customers",           _clean(customers.get("total_customers")), ""),
        ("Customers",   "Top Customer Concentration",_fmt_pct(customers.get("top_customer_concentration")), ""),
        ("Customers",   "Top 10 Concentration",      _fmt_pct(customers.get("top_10_concentration")), ""),
        ("Customers",   "Annual Retention Rate",     _fmt_pct(customers.get("customer_retention")), ""),
        ("Customers",   "Net Revenue Retention",     _fmt_pct(customers.get("net_revenue_retention")), ""),
        ("Customers",   "Avg Contract Value (ACV)",  _clean(customers.get("avg_contract_value")), ""),
    ]

    prev_cat = ""
    for i, (cat, metric, val, note) in enumerate(rows):
        r = i + 3
        alt = (i % 2 == 1)
        if cat != prev_cat:
            _subheader_cell(ws, r, 1, cat)
            prev_cat = cat
        else:
            _data_cell(ws, r, 1, "", alt_row=alt)
        _data_cell(ws, r, 2, metric, alt_row=alt)
        _data_cell(ws, r, 3, val, alt_row=alt, align="right")
        _data_cell(ws, r, 4, note, alt_row=alt)

    ws.freeze_panes = "A3"


def _build_revenue_breakdown(ws, extracted_data: dict, company_name: str):
    fin = extracted_data.get("financials", {})
    rev = fin.get("revenue", {})
    segments = fin.get("revenue_by_segment", [])
    revenue_history = rev.get("history", []) if isinstance(rev.get("history"), list) else []

    _title_row(ws, f"Revenue Breakdown — {company_name}", 4)

    # Historical trend table
    for col, header in enumerate(
        ["Period", "Revenue", "EBITDA", "EBITDA Margin"], start=1
    ):
        _header_cell(ws, 2, col, header,
                     width_hint=[22, 24, 24, 24][col - 1])

    ebitda = fin.get("ebitda", {})
    margin = fin.get("ebitda", {}).get("margin_ltm")
    hist_rows = [
        ("LTM",           rev.get("ltm"),           ebitda.get("ltm"),           ebitda.get("margin_ltm")),
        ("Prior Year",    rev.get("prior_year"),     None,                        None),
        ("Two Years Ago", rev.get("two_years_ago"),  None,                        None),
    ]
    for i, (period, r, e, m) in enumerate(hist_rows):
        row = i + 3
        alt = (i % 2 == 0)
        _data_cell(ws, row, 1, period, alt_row=alt)
        _data_cell(ws, row, 2, _fmt_num(r), alt_row=alt, align="right")
        _data_cell(ws, row, 3, _fmt_num(e), alt_row=alt, align="right")
        _data_cell(ws, row, 4, _fmt_pct(m), alt_row=alt, align="right")

    next_row = 6

    # Revenue history section (if available)
    if revenue_history:
        hist_start = next_row + 1
        ws.merge_cells(
            start_row=hist_start, start_column=1,
            end_row=hist_start, end_column=4,
        )
        cell = ws.cell(row=hist_start, column=1, value="Revenue History")
        cell.font = Font(bold=True, color=_WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(
            ["Period", "Revenue", "YoY Growth", "Notes"], start=1
        ):
            _header_cell(ws, hist_start + 1, col, header)

        for i, row in enumerate(revenue_history):
            if not isinstance(row, dict):
                continue
            r = hist_start + 2 + i
            alt = (i % 2 == 1)
            _data_cell(ws, r, 1, row.get("period", "N/A"), alt_row=alt)
            _data_cell(ws, r, 2, _fmt_num(row.get("value")), alt_row=alt, align="right")
            _data_cell(ws, r, 3, _fmt_pct(row.get("yoy_growth")), alt_row=alt, align="right")
            _data_cell(ws, r, 4, "", alt_row=alt)
        next_row = hist_start + 2 + len(revenue_history)

    # Segment table (if available)
    if segments:
        seg_start = next_row + 1
        ws.merge_cells(
            start_row=seg_start, start_column=1,
            end_row=seg_start, end_column=4,
        )
        cell = ws.cell(row=seg_start, column=1, value="Revenue by Segment")
        cell.font = Font(bold=True, color=_WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(
            ["Segment", "Revenue", "% of Total", "Notes"], start=1
        ):
            _header_cell(ws, seg_start + 1, col, header)

        for i, seg in enumerate(segments):
            r = seg_start + 2 + i
            alt = (i % 2 == 1)
            _data_cell(ws, r, 1, seg.get("segment", ""), alt_row=alt)
            _data_cell(ws, r, 2, _fmt_num(seg.get("revenue")), alt_row=alt, align="right")
            _data_cell(ws, r, 3, _fmt_pct(seg.get("pct_of_total")), alt_row=alt, align="right")
            _data_cell(ws, r, 4, "", alt_row=alt)
    else:
        ws.cell(row=next_row + 1, column=1, value="Segment data not provided in CIM.").font = Font(
            italic=True, color="888888"
        )

    ws.freeze_panes = "A3"


def _build_comp_set(ws, comps, company_name: str):
    _title_row(ws, f"Comparable Companies & Transactions — {company_name}", 7)

    headers = ["Company / Deal", "Type", "EV/Revenue", "EV/EBITDA",
               "Rationale", "Key Differences", "Confidence"]
    widths   = [28, 22, 14, 14, 50, 40, 14]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 2, col, h, width_hint=w)

    if not comps:
        ws.cell(row=3, column=1, value="No comparable companies identified.").font = Font(italic=True)
        return

    for i, comp in enumerate(comps):
        r = i + 3
        alt = (i % 2 == 1)
        _data_cell(ws, r, 1, comp.name, alt_row=alt)
        _data_cell(ws, r, 2, comp.type.replace("_", " ").title(), alt_row=alt)
        _data_cell(ws, r, 3,
                   f"{comp.ev_revenue:.1f}x" if comp.ev_revenue else "N/A",
                   alt_row=alt, align="right")
        _data_cell(ws, r, 4,
                   f"{comp.ev_ebitda:.1f}x" if comp.ev_ebitda else "N/A",
                   alt_row=alt, align="right")
        _data_cell(ws, r, 5, comp.rationale, alt_row=alt)
        _data_cell(ws, r, 6, comp.key_differences, alt_row=alt)
        conf_cell = _data_cell(ws, r, 7,
                               f"{comp.confidence:.0%}",
                               alt_row=alt, align="center")
        conf_cell.fill = _score_fill(comp.confidence)

    ws.freeze_panes = "A3"


def _build_comp_set_enhanced(
    ws,
    comps,
    company_name: str,
    extracted_data: dict,
    fund_matches: Optional[list] = None,
):
    """Comp Set sheet with standard comps + extracted peers + top fund matches."""
    _title_row(ws, f"Comparable Companies & Transactions — {company_name}", 7)

    headers = ["Company / Deal", "Type", "EV/Revenue", "EV/EBITDA",
               "Rationale", "Key Differences", "Confidence"]
    widths = [28, 22, 14, 14, 50, 40, 14]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 2, col, h, width_hint=w)

    rows = []

    # 1) Existing comparable set rows
    for comp in (comps or []):
        rows.append({
            "name": comp.name,
            "type": comp.type.replace("_", " ").title(),
            "ev_revenue": f"{comp.ev_revenue:.1f}x" if comp.ev_revenue else "N/A",
            "ev_ebitda": f"{comp.ev_ebitda:.1f}x" if comp.ev_ebitda else "N/A",
            "rationale": comp.rationale,
            "key_differences": comp.key_differences,
            "confidence": f"{comp.confidence:.0%}",
            "confidence_float": comp.confidence,
        })

    # 2) Extracted peers from document analysis (market.key_competitors)
    market = extracted_data.get("market", {})
    co = extracted_data.get("company_overview", {})
    industry = co.get("industry", "N/A")
    comp_position = market.get("competitive_position", "")

    seen_names = {r["name"].strip().lower() for r in rows if r.get("name")}
    for peer in (market.get("key_competitors") or []):
        if isinstance(peer, dict):
            name = (peer.get("name") or peer.get("company") or "").strip()
            notes = peer.get("notes") or peer.get("description") or ""
            sector = peer.get("sector") or industry
        else:
            name = str(peer).strip() if peer else ""
            notes = ""
            sector = industry
        if not name or name.lower() in seen_names or name in ("N/A", "not_provided"):
            continue
        seen_names.add(name.lower())
        rows.append({
            "name": name,
            "type": "Identified Peer",
            "ev_revenue": "N/A",
            "ev_ebitda": "N/A",
            "rationale": f"Sector: {sector}",
            "key_differences": notes or comp_position or "Extracted from company competitors list.",
            "confidence": "N/A",
            "confidence_float": None,
        })

    # 3) Top fund matches (explicitly capped at top 5)
    if fund_matches:
        for fm in fund_matches[:5]:
            thesis = getattr(fm.fund, "thesis_summary", "") or ""
            reasons = getattr(fm, "reasons", []) or []
            notes = "; ".join(str(x) for x in reasons[:2]) or "Top match from PE fund matcher."
            rows.append({
                "name": fm.fund.name,
                "type": "PE Fund Match",
                "ev_revenue": "N/A",
                "ev_ebitda": "N/A",
                "rationale": thesis,
                "key_differences": notes,
                "confidence": f"{fm.total_score:.0%}",
                "confidence_float": fm.total_score,
            })

    if not rows:
        ws.cell(row=3, column=1, value="No comparable companies identified.").font = Font(italic=True)
        return

    for i, row in enumerate(rows):
        r = i + 3
        alt = (i % 2 == 1)
        _data_cell(ws, r, 1, row["name"], alt_row=alt)
        _data_cell(ws, r, 2, row["type"], alt_row=alt)
        _data_cell(ws, r, 3, row["ev_revenue"], alt_row=alt, align="right")
        _data_cell(ws, r, 4, row["ev_ebitda"], alt_row=alt, align="right")
        _data_cell(ws, r, 5, row["rationale"], alt_row=alt)
        _data_cell(ws, r, 6, row["key_differences"], alt_row=alt)
        conf_cell = _data_cell(ws, r, 7, row["confidence"], alt_row=alt, align="center")
        if isinstance(row.get("confidence_float"), (int, float)):
            conf_cell.fill = _score_fill(float(row["confidence_float"]))

    ws.freeze_panes = "A3"


def _extract_memo_risks_for_export(memo_text: str) -> list:
    """Extract qualitative memo risks from JSON memo (sections format) or markdown."""
    if not memo_text:
        return []

    def _strip_md(text: str) -> str:
        text = re.sub(r'\*\*(.+?)\*\*', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'\*(.+?)\*',     r'\1', text, flags=re.DOTALL)
        return text.strip()

    def _mk_row(risk: str, mitigant: str = "", description: str = "") -> Optional[dict]:
        risk        = _strip_md((risk        or "").strip())
        mitigant    = _strip_md((mitigant    or "").strip())
        description = _strip_md((description or "").strip())
        if not risk or len(risk) < 5:
            return None
        title = risk.split(":", 1)[0].strip()[:120]
        return {
            "category":           "Memo",
            "severity":           "Medium",
            "title":              title,
            "description":        description or risk,
            "mitigant":           mitigant or "—",
            "diligence_question": "—",
        }

    def _parse_risk_text(text: str) -> list:
        """Parse numbered risk blocks with optional Mitigant: lines."""
        results = []
        blocks = re.split(r'\n(?=\d+[\.\)])', text.strip())
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            if not re.match(r'^\d+', block):
                if len(block) > 20:
                    results.append(_mk_row(block))
                continue
            parts = re.split(
                r'\n?\s*(?:\*\*?)?Mitigant(?:\*\*?)?\s*:',
                block, maxsplit=1, flags=re.IGNORECASE,
            )
            risk_raw = re.sub(r'^\d+[\.\)]\s*', '', parts[0]).strip()
            mitigant = parts[1].strip() if len(parts) > 1 else ""
            row = _mk_row(risk_raw, mitigant)
            if row:
                results.append(row)
        return [r for r in results if r]

    out = []

    # ── JSON parsing ───────────────────────────────────────────────────────
    try:
        data = json.loads(memo_text)
    except (json.JSONDecodeError, TypeError):
        data = None

    if isinstance(data, dict):
        # Path 1: flat sections array (standard memo format)
        for section in (data.get("sections") or []):
            heading = (section.get("heading") or section.get("title") or "").lower()
            if not ("risk" in heading or "mitigant" in heading):
                continue
            content = section.get("content") or ""
            if isinstance(content, list):
                for item in content:
                    if isinstance(item, dict):
                        row = _mk_row(
                            item.get("risk")     or item.get("title")      or "",
                            item.get("mitigant") or item.get("mitigation") or "",
                            item.get("description") or "",
                        )
                        if row:
                            out.append(row)
            elif isinstance(content, str) and content.strip():
                out.extend(_parse_risk_text(content))
            if out:
                return out

        # Path 2: nested investment_committee_memo structure
        icm = data.get("investment_committee_memo") or {}
        risk_list = icm.get("key_risks_and_mitigants") or []
        if isinstance(risk_list, list):
            for item in risk_list:
                if isinstance(item, dict):
                    row = _mk_row(
                        item.get("risk")     or item.get("title")      or "",
                        item.get("mitigant") or item.get("mitigation") or "",
                        item.get("description") or "",
                    )
                    if row:
                        out.append(row)
            if out:
                return out

    # ── Markdown fallback ──────────────────────────────────────────────────
    section_match = re.search(
        r"(?:KEY RISKS|RISKS?\s*&\s*MITIGANTS?|key_risks_and_mitigants)(.*?)"
        r"(?:\n#+\s|\n\d+\.\s+[A-Z][A-Z]|\Z)",
        memo_text,
        re.DOTALL | re.IGNORECASE,
    )
    if section_match:
        out.extend(_parse_risk_text(section_match.group(1).strip()))

    return out


def _build_risk_register(ws, risks, company_name: str, memo_text: str = ""):
    _title_row(ws, f"Risk Register — {company_name}", 6)

    headers = ["Category", "Severity", "Risk Title",
               "Description", "Mitigant", "Diligence Question"]
    widths   = [18, 12, 30, 50, 40, 50]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 2, col, h, width_hint=w)

    all_rows = []

    for risk in (risks or []):
        all_rows.append({
            "category": risk.category,
            "severity": risk.severity,
            "title": risk.title,
            "description": risk.description,
            "mitigant": risk.mitigant or "—",
            "diligence_question": risk.diligence_question or "—",
        })

    all_rows.extend(_extract_memo_risks_for_export(memo_text))

    if not all_rows:
        ws.cell(row=3, column=1, value="No risks identified.").font = Font(italic=True)
        return

    severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    sorted_risks = sorted(all_rows, key=lambda r: severity_order.get(r.get("severity"), 4))

    for i, risk in enumerate(sorted_risks):
        r = i + 3
        alt = (i % 2 == 1)
        _data_cell(ws, r, 1, risk.get("category", "General"), alt_row=alt)

        # Severity cell with background colour
        sev = risk.get("severity", "Medium")
        sev_cell = ws.cell(row=r, column=2, value=sev)
        sev_cell.fill = _severity_fill(sev)
        sev_cell.font = Font(bold=True, color=_WHITE, size=10)
        sev_cell.alignment = Alignment(horizontal="center", vertical="center")
        sev_cell.border = _thin_border()

        _data_cell(ws, r, 3, risk.get("title", ""), alt_row=alt)
        _data_cell(ws, r, 4, risk.get("description", ""), alt_row=alt)
        _data_cell(ws, r, 5, risk.get("mitigant", "—"), alt_row=alt)
        _data_cell(ws, r, 6, risk.get("diligence_question", "—"), alt_row=alt)
        ws.row_dimensions[r].height = 40

    ws.freeze_panes = "A3"


def _build_deal_score(ws, deal_score, company_name: str):
    if deal_score is None:
        ws.cell(row=1, column=1, value="Deal score not computed.")
        return

    _title_row(ws, f"Deal Score — {company_name}", 6)

    # Summary block (rows 2-5)
    summary_data = [
        ("Overall Score",    f"{deal_score.total_score:.0%}"),
        ("Grade",            deal_score.grade),
        ("Recommendation",   deal_score.recommendation),
        ("Summary",          deal_score.summary),
    ]
    for i, (label, val) in enumerate(summary_data):
        r = i + 2
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_l.font = Font(bold=True, color=_NAVY, size=11)
        cell_l.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
        cell_l.alignment = Alignment(horizontal="left", vertical="center")
        cell_l.border = _thin_border()

        ws.merge_cells(
            start_row=r, start_column=2,
            end_row=r, end_column=6,
        )
        cell_v = ws.cell(row=r, column=2, value=val)
        cell_v.font = Font(size=11, bold=(label == "Overall Score"))
        if label == "Overall Score":
            cell_v.fill = _score_fill(deal_score.total_score)
        cell_v.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        cell_v.border = _thin_border()
        ws.row_dimensions[r].height = 20 if label != "Summary" else 50

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # Dimension breakdown table
    dim_start = 7
    headers = ["Dimension", "Score", "Weight", "Weighted Score", "Data Quality", "Rationale"]
    widths   = [28, 12, 12, 16, 14, 60]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, dim_start, col, h, width_hint=w)

    for i, dim in enumerate(deal_score.dimensions):
        r = dim_start + 1 + i
        alt = (i % 2 == 1)

        _data_cell(ws, r, 1, dim.dimension, alt_row=alt)

        score_cell = ws.cell(row=r, column=2, value=f"{dim.score:.0%}")
        score_cell.fill = _score_fill(dim.score)
        score_cell.font = Font(bold=True, color=_WHITE, size=10)
        score_cell.alignment = Alignment(horizontal="center", vertical="center")
        score_cell.border = _thin_border()

        _data_cell(ws, r, 3, f"{dim.weight:.0%}", alt_row=alt, align="center")
        _data_cell(ws, r, 4, f"{dim.weighted_score:.3f}", alt_row=alt, align="center")

        dq_cell = ws.cell(row=r, column=5, value=dim.data_quality.replace("_", " ").title())
        dq_colours = {"Strong": _GREEN, "Moderate": _AMBER,
                      "Weak": _ORANGE, "No Data": _MED_GREY}
        dq_cell.fill = PatternFill(
            "solid", fgColor=dq_colours.get(
                dim.data_quality.replace("_", " ").title(), _MED_GREY
            )
        )
        dq_cell.alignment = Alignment(horizontal="center", vertical="center")
        dq_cell.border = _thin_border()

        _data_cell(ws, r, 6, dim.rationale, alt_row=alt)
        ws.row_dimensions[r].height = 35

    ws.freeze_panes = "A8"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _build_narrative_validation(ws, narrative_gaps: list, company_name: str):
    """Narrative Validation (banker spin detector) sheet."""
    _title_row(ws, f"Narrative Validation — {company_name}", 6)

    headers = ["Claim", "Source", "Extracted Value", "Status", "Gap / Delta", "Severity"]
    widths  = [40, 28, 30, 14, 55, 12]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 2, col, h, width_hint=w)

    _status_colors = {
        "confirmed":    "70AD47",   # green
        "discrepancy":  "FF0000",   # red
        "unverifiable": "FFC000",   # amber
    }
    _sev_colors = {
        "critical": "C00000",
        "warning":  "FFC000",
        "info":     "70AD47",
    }

    if not narrative_gaps:
        ws.cell(row=3, column=1,
                value="No narrative gaps detected (memo not generated or no checkable claims found)."
                ).font = Font(italic=True, color="888888")
        return

    # Sort: discrepancies first, then unverifiable, then confirmed
    order = {"discrepancy": 0, "unverifiable": 1, "confirmed": 2}
    sev_order = {"critical": 0, "warning": 1, "info": 2}
    sorted_gaps = sorted(
        narrative_gaps,
        key=lambda g: (order.get(g.get("status", ""), 9),
                       sev_order.get(g.get("severity", ""), 9)),
    )

    for i, gap in enumerate(sorted_gaps):
        r   = i + 3
        alt = (i % 2 == 1)
        status   = gap.get("status", "")
        severity = gap.get("severity", "")

        _data_cell(ws, r, 1, gap.get("claim", ""),          alt_row=alt)
        _data_cell(ws, r, 2, gap.get("claim_source", ""),   alt_row=alt)
        _data_cell(ws, r, 3, gap.get("extracted_value", ""),alt_row=alt)

        # Status cell with colour
        stat_cell = ws.cell(row=r, column=4, value=status.title())
        stat_cell.fill  = PatternFill("solid", fgColor=_status_colors.get(status, _MED_GREY))
        stat_cell.font  = Font(bold=True, color=_WHITE, size=10)
        stat_cell.alignment = Alignment(horizontal="center", vertical="center")
        stat_cell.border = _thin_border()

        _data_cell(ws, r, 5, gap.get("gap") or "—", alt_row=alt)

        sev_cell = ws.cell(row=r, column=6, value=severity.title())
        sev_cell.fill  = PatternFill("solid", fgColor=_sev_colors.get(severity, _MED_GREY))
        sev_cell.font  = Font(bold=True, color=_WHITE if severity != "warning" else "111111", size=10)
        sev_cell.alignment = Alignment(horizontal="center", vertical="center")
        sev_cell.border = _thin_border()

        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A3"


def export_excel(result: "AnalysisResult", output_path: str) -> str:
    """Build and save the Excel workbook from an AnalysisResult.

    Args:
        result: Full AnalysisResult from MeridianPipeline.analyze().
        output_path: Destination .xlsx file path.

    Returns:
        Absolute path to the saved file.
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    co = result.extracted_data.get("company_overview", {})
    company_name = co.get("company_name", "Unknown Company")

    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # --- Tab 1: Financial Summary ---
    ws1 = wb.create_sheet("Financial Summary")
    _build_financial_summary(ws1, result.extracted_data, company_name)

    # --- Tab 2: Revenue Breakdown ---
    ws2 = wb.create_sheet("Revenue Breakdown")
    _build_revenue_breakdown(ws2, result.extracted_data, company_name)

    # --- Tab 3: Comp Set ---
    ws3 = wb.create_sheet("Comp Set")
    fund_matches = getattr(result, "fund_matches", None)
    if not fund_matches:
        try:
            from scoring.fund_matcher import MatchingEngine
            fund_matches = MatchingEngine().match(result.extracted_data, top_n=5)
        except Exception:
            fund_matches = []
    _build_comp_set_enhanced(
        ws3,
        result.comps or [],
        company_name,
        result.extracted_data,
        fund_matches=fund_matches,
    )

    # --- Tab 4: Risk Register ---
    ws4 = wb.create_sheet("Risk Register")
    _build_risk_register(ws4, result.risks or [], company_name, memo_text=result.memo or "")

    # --- Tab 5: Deal Score ---
    ws5 = wb.create_sheet("Deal Score")
    _build_deal_score(ws5, result.deal_score, company_name)

    # --- Tab 6: Narrative Validation (banker spin detector) ---
    narrative_gaps = getattr(result, "narrative_gaps", None) or []
    ws6 = wb.create_sheet("Narrative Validation")
    _build_narrative_validation(ws6, narrative_gaps, company_name)

    # Set tab colours
    ws1.sheet_properties.tabColor = _NAVY
    ws2.sheet_properties.tabColor = _DARK_BLUE
    ws3.sheet_properties.tabColor = "4472C4"
    ws4.sheet_properties.tabColor = _DARK_RED
    ws5.sheet_properties.tabColor = _GREEN
    ws6.sheet_properties.tabColor = _ORANGE

    wb.save(output_path)
    return os.path.abspath(output_path)


def export_batch_comparison(results: list, output_path: str) -> str:
    """Create a side-by-side comparison workbook for batch analysis.

    Args:
        results: List of (filename, AnalysisResult) tuples.
        output_path: Destination .xlsx path.

    Returns:
        Absolute path to saved file.
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Comparison"

    headers = [
        "File", "Company", "Industry", "Sub-Industry", "Business Model",
        "Headquarters", "Revenue LTM ($M)", "EBITDA LTM ($M)", "EBITDA Margin",
        "Revenue CAGR (3yr)", "Recurring Revenue %", "Top Cust. Conc.",
        "Employees", "Total Score", "Grade", "Recommendation",
        "Critical Risks", "High Risks", "# Comps",
    ]
    widths = [
        25, 25, 20, 20, 16, 20, 18, 18, 16, 18, 20, 18, 12,
        14, 8, 30, 14, 10, 10,
    ]

    _title_row(ws, f"Meridian AI — Batch Deal Comparison ({datetime.utcnow():%Y-%m-%d})",
               len(headers))
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 2, col, h, width_hint=w)

    for i, (filename, result) in enumerate(results):
        r = i + 3
        alt = (i % 2 == 1)
        co  = result.extracted_data.get("company_overview", {})
        fin = result.extracted_data.get("financials", {})
        rev = fin.get("revenue", {})
        ebitda = fin.get("ebitda", {})
        customers = result.extracted_data.get("customers", {})
        ds = result.deal_score

        critical_risks = len([x for x in (result.risks or []) if x.severity == "Critical"])
        high_risks     = len([x for x in (result.risks or []) if x.severity == "High"])

        row_vals = [
            os.path.basename(filename),
            co.get("company_name", "Unknown"),
            co.get("industry", "N/A"),
            co.get("sub_industry", "N/A"),
            co.get("business_model", "N/A"),
            co.get("headquarters", "N/A"),
            _to_m_float(rev.get("ltm")),
            _to_m_float(ebitda.get("ltm")),
            _fmt_pct(ebitda.get("margin_ltm")),
            _fmt_pct(rev.get("cagr_3yr")),
            _fmt_pct(fin.get("recurring_revenue_pct")),
            _fmt_pct(customers.get("top_customer_concentration")),
            co.get("employees", "N/A"),
            f"{ds.total_score:.0%}" if ds else "N/A",
            ds.grade if ds else "N/A",
            ds.recommendation if ds else "N/A",
            critical_risks,
            high_risks,
            len(result.comps or []),
        ]

        for col, val in enumerate(row_vals, start=1):
            cell = _data_cell(ws, r, col, val, alt_row=alt,
                              align="right" if isinstance(val, (int, float)) else "left")

        # Colour-code the score and grade cells
        if ds:
            score_col = headers.index("Total Score") + 1
            grade_col = headers.index("Grade") + 1
            ws.cell(row=r, column=score_col).fill = _score_fill(ds.total_score)
            ws.cell(row=r, column=grade_col).fill = _score_fill(ds.total_score)

    ws.freeze_panes = "A3"
    wb.save(output_path)
    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _clean(val) -> str:
    if val is None or val == "not_provided":
        return "N/A"
    return str(val)


def _fmt_num(val) -> str:
    if val is None or val == "not_provided":
        return "N/A"
    try:
        v = float(val)
        sign = "-" if v < 0 else ""
        abs_v = abs(v)
        if abs_v >= 1_000_000:
            return f"{sign}${abs_v / 1_000_000:.1f}M"
        if abs_v >= 1_000:
            return f"{sign}${abs_v / 1_000:.1f}K"
        return f"{sign}${abs_v:,.0f}"
    except (ValueError, TypeError):
        return str(val)


def _fmt_pct(val) -> str:
    if val is None or val == "not_provided":
        return "N/A"
    try:
        return f"{float(val):.1%}"
    except (ValueError, TypeError):
        return str(val)


def _to_m_float(val) -> Optional[float]:
    if val is None or val == "not_provided":
        return None
    try:
        v = float(val)
        return round(v / 1_000_000, 2) if v >= 1_000_000 else round(v, 2)
    except (ValueError, TypeError):
        return None
